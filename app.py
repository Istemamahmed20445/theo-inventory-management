from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, send_file
from flask_cors import CORS
import firebase_admin
from firebase_admin import credentials, firestore, storage
import os
import re
import unicodedata
from functools import wraps
import time
import logging
from config import config

def secure_filename(filename):
    """Secure a filename for storage."""
    if filename is None:
        return None
    filename = unicodedata.normalize('NFKD', filename)
    filename = filename.encode('ascii', 'ignore').decode('ascii')
    filename = re.sub(r'[^\w\s-]', '', filename)
    filename = re.sub(r'[-\s]+', '-', filename)
    return filename.strip('-')
from datetime import datetime
import json
import uuid
import qrcode
from PIL import Image
import io
import base64
import openpyxl
from openpyxl import Workbook
import io
from functools import wraps

app = Flask(__name__)

# Load configuration based on environment
config_name = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config[config_name])

# Set secret key from environment or use default
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
CORS(app)

# Template helper function for permission checking
@app.template_global()
def has_permission(permission):
    """Check if current user has a specific permission"""
    return permission in session.get('permissions', [])

# Firebase availability check
def check_firebase():
    """Check if Firebase is available"""
    if db is None:
        flash('Firebase is not available. Please check your configuration.', 'error')
        return False
    return True

# Performance monitoring decorator
def performance_monitor(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        start_time = time.time()
        result = f(*args, **kwargs)
        end_time = time.time()
        print(f"Function {f.__name__} took {end_time - start_time:.4f} seconds")
        return result
    return decorated_function

# Enhanced cache for frequently accessed data
cache = {
    'products': {'data': None, 'timestamp': 0},
    'categories': {'data': None, 'timestamp': 0},
    'sizes': {'data': None, 'timestamp': 0},
    'colors': {'data': None, 'timestamp': 0},
    'customers': {'data': None, 'timestamp': 0},
    'sales_orders': {'data': None, 'timestamp': 0},
    'production_orders': {'data': None, 'timestamp': 0},
    'users': {'data': None, 'timestamp': 0},
    'dashboard_stats': {'data': None, 'timestamp': 0},
    'recent_activities': {'data': None, 'timestamp': 0}
}

CACHE_DURATION = 300  # 5 minutes for better performance
QUICK_CACHE_DURATION = 60  # 1 minute for frequently changing data

# Helper function to get cached data
def get_cached_data(cache_key, fetch_function, duration=CACHE_DURATION):
    current_time = time.time()
    if (cache[cache_key]['data'] is None or 
        current_time - cache[cache_key]['timestamp'] > duration):
        cache[cache_key]['data'] = fetch_function()
        cache[cache_key]['timestamp'] = current_time
    return cache[cache_key]['data']

# Optimized database query helper
def get_products_optimized(limit=None, order_by=None):
    """Optimized product fetching with optional limit and ordering"""
    products_ref = db.collection('products')
    
    if order_by:
        products_ref = products_ref.order_by(order_by)
    
    if limit:
        products_ref = products_ref.limit(limit)
    
    products = products_ref.get()
    return [{'id': product.id, **product.to_dict()} for product in products]

# Optimized statistics calculation
def calculate_dashboard_stats():
    """Calculate dashboard statistics efficiently"""
    def fetch_stats():
        products = get_products_optimized()
        return {
            'total_products': len(products),
            'total_value': sum(p.get('price', 0) for p in products)
        }
    
    return get_cached_data('dashboard_stats', fetch_stats, QUICK_CACHE_DURATION)

def parse_item_numbers(item_numbers):
    """Parse item numbers string to get count (e.g., '1,2,3' or '1-5' or '1,3-5,7' or '40')"""
    if not item_numbers:
        return 0
    
    total_items = 0
    parts = item_numbers.split(',')
    
    for part in parts:
        part = part.strip()
        if '-' in part:
            # Handle ranges like "1-5"
            try:
                start, end = part.split('-')
                start, end = int(start.strip()), int(end.strip())
                if start <= end:
                    total_items += (end - start + 1)
            except ValueError:
                pass
        else:
            # Handle single numbers - treat as quantity if it's a simple number
            try:
                num = int(part)
                # If it's just a number without other characters, treat as quantity
                if part == str(num):
                    total_items += num
                else:
                    # If it contains other characters, treat as single item
                    total_items += 1
            except ValueError:
                pass
    
    return total_items

# Firebase configuration
# Prefer service account JSON from environment on Render; fallback to local file or ADC
service_account_json = os.environ.get('FIREBASE_SERVICE_ACCOUNT_JSON')

firebase_credentials = None
if service_account_json:
    try:
        # Decode base64 if it's base64 encoded
        if service_account_json.startswith('ewog'):  # Base64 encoded JSON starts with 'ewog'
            import base64
            decoded_json = base64.b64decode(service_account_json).decode('utf-8')
            service_account_info = json.loads(decoded_json)
        else:
            # Try parsing as direct JSON
            service_account_info = json.loads(service_account_json)
        firebase_credentials = credentials.Certificate(service_account_info)
        print("Firebase credentials loaded from environment variable")
    except Exception as e:
        print(f"Error loading Firebase credentials from environment: {e}")
        firebase_credentials = None

if firebase_credentials is None:
    service_account_file = os.environ.get('FIREBASE_CREDENTIALS_FILE', 'firebase-service-account.json')
    if os.path.exists(service_account_file):
        try:
            firebase_credentials = credentials.Certificate(service_account_file)
            print("Firebase credentials loaded from file")
        except Exception as e:
            print(f"Error loading Firebase credentials from file: {e}")
            firebase_credentials = None
    else:
        print("No Firebase credentials found")
        firebase_credentials = None

firebase_options = {}
env_storage_bucket = os.environ.get('FIREBASE_STORAGE_BUCKET')
config_storage_bucket = app.config.get('FIREBASE_STORAGE_BUCKET')
resolved_storage_bucket = env_storage_bucket or config_storage_bucket
if resolved_storage_bucket:
    firebase_options['storageBucket'] = resolved_storage_bucket

# Initialize Firebase only if credentials are available
if firebase_credentials:
    try:
        firebase_admin.initialize_app(firebase_credentials, firebase_options or None)
        db = firestore.client()
        bucket = storage.bucket()
        print("Firebase initialized successfully")
    except Exception as e:
        print(f"Error initializing Firebase: {e}")
        db = None
        bucket = None
else:
    print("Firebase credentials not available - running without Firebase")
    db = None
    bucket = None

# Configuration
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Create default admin user if no users exist
def create_default_admin():
    try:
        users_ref = db.collection('users')
        existing_users = users_ref.get()
        
        if not existing_users:
            # Create default admin user
            admin_data = {
                'username': 'admin',
                'password': 'admin123',
                'role': 'admin',
                'permissions': ['view_products', 'add_products', 'edit_products', 'delete_products', 'excel_import', 'view_reports', 'sales_customer'],
                'created_at': datetime.now(),
                'active': True
            }
            db.collection('users').add(admin_data)
            print("THEO CLOTHING INVENTORY - Default admin user created:")
            print("Username: admin")
            print("Password: admin123")
    except Exception as e:
        print(f"Error creating default admin: {e}")

# Create default admin on startup
create_default_admin()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user_doc = db.collection('users').document(session['user_id']).get()
        if not user_doc.exists or user_doc.to_dict().get('role') != 'admin':
            flash('Admin access required', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            
            user_permissions = session.get('permissions', [])
            if permission not in user_permissions:
                flash(f'Permission denied: {permission} required', 'error')
                return redirect(url_for('dashboard'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/health')
def health_check():
    """Health check endpoint for load balancers and monitoring"""
    try:
        # Check if Firebase connection is working
        db.collection('health_check').limit(1).get()
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now().isoformat(),
            'version': '1.0.0'
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 503

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Check user credentials
        users_ref = db.collection('users')
        query = users_ref.where('username', '==', username).where('password', '==', password)
        users = query.get()
        
        if users:
            user = users[0]
            user_data = user.to_dict()
            session['user_id'] = user.id
            session['username'] = user_data['username']
            session['role'] = user_data['role']
            session['permissions'] = user_data.get('permissions', [])
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
@performance_monitor
def dashboard():
    # Get recent activities (cached for 1 minute for better performance)
    def fetch_activities():
        activities_ref = db.collection('activities').order_by('timestamp', direction=firestore.Query.DESCENDING).limit(10)
        return [doc.to_dict() for doc in activities_ref.get()]
    
    # Use optimized statistics calculation
    stats = calculate_dashboard_stats()
    
    # Get recent activities with quick cache
    recent_activities = get_cached_data('recent_activities', fetch_activities, QUICK_CACHE_DURATION)
    
    return render_template('dashboard.html', 
                         total_products=stats['total_products'],
                         total_value=stats['total_value'],
                         recent_activities=recent_activities)

@app.route('/products')
@login_required
@performance_monitor
def products():
    # Get pagination parameters
    page = int(request.args.get('page', 1))
    per_page = 20  # Show 20 products per page
    
    def fetch_products():
        return get_products_optimized()
    
    # Get categories for filter dropdown
    def fetch_categories():
        categories_ref = db.collection('categories')
        categories = categories_ref.get()
        category_list = []
        
        for category in categories:
            category_data = category.to_dict()
            category_data['id'] = category.id
            category_list.append(category_data)
        
        return category_list
    
    all_products = get_cached_data('products', fetch_products)
    categories = get_cached_data('categories', fetch_categories)
    
    # Calculate pagination
    total_products = len(all_products)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    product_list = all_products[start_idx:end_idx]
    
    # Calculate pagination info
    total_pages = (total_products + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    
    return render_template('products.html', 
                         products=product_list,
                         categories=categories,
                         page=page,
                         total_pages=total_pages,
                         has_prev=has_prev,
                         has_next=has_next,
                         total_products=total_products)

@app.route('/add_product', methods=['GET', 'POST'])
@login_required
@permission_required('add_products')
def add_product():
    # Get categories for dropdown
    def fetch_categories():
        categories_ref = db.collection('categories')
        categories = categories_ref.get()
        category_list = []
        
        for category in categories:
            category_data = category.to_dict()
            category_data['id'] = category.id
            category_list.append(category_data)
        
        return category_list
    
    categories = get_cached_data('categories', fetch_categories)
    
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form['name']
            category = request.form['category']
            size = request.form.get('size', '')
            color = request.form.get('color', '')
            price = float(request.form['price'])
            body_size = request.form.get('body_size', '')
            waist_size = request.form.get('waist_size', '')
            length = request.form.get('length', '')
            description = request.form.get('description', '')
            
            # Handle image upload
            image_url = ''
            if 'image' in request.files:
                file = request.files['image']
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    file.save(file_path)
                    
                    # Upload to Firebase Storage
                    blob = bucket.blob(f"product_images/{unique_filename}")
                    blob.upload_from_filename(file_path)
                    blob.make_public()
                    image_url = blob.public_url
                    
                    # Delete local file
                    os.remove(file_path)
            
            # Generate barcode
            barcode_data = f"PROD_{uuid.uuid4().hex[:12].upper()}"
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(barcode_data)
            qr.make(fit=True)
            
            # Create QR code image
            qr_img = qr.make_image(fill_color="black", back_color="white")
            qr_buffer = io.BytesIO()
            qr_img.save(qr_buffer, format='PNG')
            qr_buffer.seek(0)
            
            # Upload QR code to Firebase Storage
            qr_blob = bucket.blob(f"barcodes/{barcode_data}.png")
            qr_blob.upload_from_file(qr_buffer, content_type='image/png')
            qr_blob.make_public()
            qr_url = qr_blob.public_url
            
            # Save product to Firestore
            product_data = {
                'name': name,
                'category': category,
                'size': size,
                'color': color,
                'price': price,
                'body_size': body_size,
                'waist_size': waist_size,
                'length': length,
                'description': description,
                'image_url': image_url,
                'barcode': barcode_data,
                'qr_code_url': qr_url,
                'created_at': datetime.now(),
                'updated_at': datetime.now()
            }
            
            doc_ref = db.collection('products').add(product_data)
            
            # Invalidate cache
            cache['products']['data'] = None
            
            # Log activity
            activity_data = {
                'action': 'Product Added',
                'details': f'Added product: {name}',
                'user': session['username'],
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
            
            flash('Product added successfully!', 'success')
            return redirect(url_for('products'))
            
        except Exception as e:
            flash(f'Error adding product: {str(e)}', 'error')
    
    return render_template('add_product.html', categories=categories)

@app.route('/edit_product/<product_id>', methods=['GET', 'POST'])
@login_required
@permission_required('edit_products')
def edit_product(product_id):
    product_ref = db.collection('products').document(product_id)
    product_doc = product_ref.get()
    
    if not product_doc.exists:
        flash('Product not found', 'error')
        return redirect(url_for('products'))
    
    if request.method == 'POST':
        try:
            # Get form data
            name = request.form['name']
            category = request.form['category']
            size = request.form.get('size', '')
            color = request.form.get('color', '')
            price = float(request.form['price'])
            body_size = request.form.get('body_size', '')
            waist_size = request.form.get('waist_size', '')
            length = request.form.get('length', '')
            description = request.form.get('description', '')
            
            # Handle image upload
            image_url = product_doc.to_dict().get('image_url', '')
            if 'image' in request.files:
                file = request.files['image']
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    file.save(file_path)
                    
                    # Upload to Firebase Storage
                    blob = bucket.blob(f"product_images/{unique_filename}")
                    blob.upload_from_filename(file_path)
                    blob.make_public()
                    image_url = blob.public_url
                    
                    # Delete local file
                    os.remove(file_path)
            
            # Update product
            product_data = {
                'name': name,
                'category': category,
                'size': size,
                'color': color,
                'price': price,
                'body_size': body_size,
                'waist_size': waist_size,
                'length': length,
                'description': description,
                'image_url': image_url,
                'updated_at': datetime.now()
            }
            
            product_ref.update(product_data)
            
            # Invalidate cache
            cache['products']['data'] = None
            
            # Log activity
            activity_data = {
                'action': 'Product Updated',
                'details': f'Updated product: {name}',
                'user': session['username'],
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
            
            flash('Product updated successfully!', 'success')
            return redirect(url_for('products'))
            
        except Exception as e:
            flash(f'Error updating product: {str(e)}', 'error')
    
    # Get categories for dropdown
    def fetch_categories():
        categories_ref = db.collection('categories')
        categories = categories_ref.get()
        category_list = []
        
        for category in categories:
            category_data = category.to_dict()
            category_data['id'] = category.id
            category_list.append(category_data)
        
        return category_list
    
    product_data = product_doc.to_dict()
    product_data['id'] = product_id
    categories = get_cached_data('categories', fetch_categories)
    
    return render_template('edit_product.html', product=product_data, categories=categories)

@app.route('/lookup_product_by_barcode/<barcode>')
@login_required
def lookup_product_by_barcode(barcode):
    """Look up product by barcode"""
    try:
        products_ref = db.collection('products')
        product_query = products_ref.where('barcode', '==', barcode).get()
        
        if product_query:
            product_data = product_query[0].to_dict()
            product_data['id'] = product_query[0].id
            return jsonify({
                'success': True,
                'product': product_data
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Product not found'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error looking up product: {str(e)}'
        })

@app.route('/delete_product/<product_id>')
@login_required
@permission_required('delete_products')
def delete_product(product_id):
    try:
        product_ref = db.collection('products').document(product_id)
        product_doc = product_ref.get()
        
        if product_doc.exists:
            product_data = product_doc.to_dict()
            product_ref.delete()
            
            # Invalidate cache
            cache['products']['data'] = None
            
            # Log activity
            activity_data = {
                'action': 'Product Deleted',
                'details': f'Deleted product: {product_data.get("name", "Unknown")}',
                'user': session['username'],
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
            
            flash('Product deleted successfully!', 'success')
        else:
            flash('Product not found', 'error')
            
    except Exception as e:
        flash(f'Error deleting product: {str(e)}', 'error')
    
    return redirect(url_for('products'))

@app.route('/admin')
@admin_required
def admin_panel():
    # Get all users
    users_ref = db.collection('users')
    users = users_ref.get()
    user_list = []
    
    for user in users:
        user_data = user.to_dict()
        user_data['id'] = user.id
        # Debug: Log user data
        print(f"DEBUG: User {user_data.get('username')} permissions: {user_data.get('permissions')}")
        user_list.append(user_data)
    
    return render_template('admin.html', users=user_list)

@app.route('/create_user', methods=['POST'])
@admin_required
def create_user():
    try:
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        permissions = request.form.getlist('permissions')
        
        # Check if username already exists
        users_ref = db.collection('users')
        existing_user = users_ref.where('username', '==', username).get()
        
        if existing_user:
            flash('Username already exists', 'error')
            return redirect(url_for('admin_panel'))
        
        # Create new user
        user_data = {
            'username': username,
            'password': password,
            'role': role,
            'permissions': permissions,
            'created_at': datetime.now(),
            'active': True
        }
        
        db.collection('users').add(user_data)
        
        # Log activity
        activity_data = {
            'action': 'User Created',
            'details': f'Created user: {username} with role: {role}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        flash('User created successfully!', 'success')
        
    except Exception as e:
        flash(f'Error creating user: {str(e)}', 'error')
    
    return redirect(url_for('admin_panel'))

@app.route('/admin/update_user', methods=['POST'])
@admin_required
def update_user():
    try:
        user_id = request.form['user_id']
        username = request.form['username']
        password = request.form.get('password', '')
        role = request.form['role']
        permissions = request.form.getlist('permissions')
        active = request.form.get('active') == 'on'
        
        # Debug: Log received data
        print(f"DEBUG: Updating user {user_id}")
        print(f"DEBUG: Username: {username}")
        print(f"DEBUG: Role: {role}")
        print(f"DEBUG: Permissions: {permissions}")
        print(f"DEBUG: Active: {active}")
        
        # Get user document
        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            return jsonify({'success': False, 'message': 'User not found'})
        
        # Prepare update data
        update_data = {
            'role': role,
            'permissions': permissions,
            'active': active,
            'updated_at': datetime.now(),
            'updated_by': session['username']
        }
        
        # Only update password if provided
        if password:
            update_data['password'] = password
        
        print(f"DEBUG: Update data: {update_data}")
        
        # Update user
        user_ref.update(update_data)
        
        # Log activity
        activity_data = {
            'action': 'User Updated',
            'details': f'Updated user: {username} with permissions: {permissions}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'User updated successfully'})
        
    except Exception as e:
        print(f"DEBUG: Error updating user: {str(e)}")
        return jsonify({'success': False, 'message': f'Error updating user: {str(e)}'})

@app.route('/admin/generate_missing_qr_codes', methods=['POST'])
@admin_required
def generate_missing_qr_codes():
    """Generate QR codes for products that don't have them"""
    try:
        # Get all products
        products_ref = db.collection('products')
        products = products_ref.get()
        
        updated_count = 0
        bucket = storage.bucket()
        
        for product in products:
            product_data = product.to_dict()
            product_id = product.id
            
            # Check if product already has QR code
            if not product_data.get('barcode') or not product_data.get('qr_code_url'):
                # Generate barcode
                barcode_data = f"PROD_{uuid.uuid4().hex[:12].upper()}"
                qr = qrcode.QRCode(version=1, box_size=10, border=5)
                qr.add_data(barcode_data)
                qr.make(fit=True)
                
                # Create QR code image
                qr_img = qr.make_image(fill_color="black", back_color="white")
                qr_buffer = io.BytesIO()
                qr_img.save(qr_buffer, format='PNG')
                qr_buffer.seek(0)
                
                # Upload QR code to Firebase Storage
                qr_blob = bucket.blob(f"barcodes/{barcode_data}.png")
                qr_blob.upload_from_file(qr_buffer, content_type='image/png')
                qr_blob.make_public()
                qr_url = qr_blob.public_url
                
                # Update product with barcode and QR code
                product.reference.update({
                    'barcode': barcode_data,
                    'qr_code_url': qr_url,
                    'updated_at': datetime.now()
                })
                
                updated_count += 1
        
        # Log activity
        activity_data = {
            'action': 'Bulk QR Generation',
            'details': f'Generated QR codes for {updated_count} products',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({
            'success': True,
            'message': f'Generated QR codes for {updated_count} products'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generating QR codes: {str(e)}'
        })

@app.route('/admin/delete_user', methods=['POST'])
@admin_required
def delete_user():
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({'success': False, 'message': 'User ID is required'})
        
        # Get user document
        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            return jsonify({'success': False, 'message': 'User not found'})
        
        user_data = user_doc.to_dict()
        username = user_data.get('username', 'Unknown')
        
        # Prevent admin from deleting themselves
        if user_id == session.get('user_id'):
            return jsonify({'success': False, 'message': 'You cannot delete your own account'})
        
        # Prevent deletion of the last admin
        if user_data.get('role') == 'admin':
            admin_users = db.collection('users').where('role', '==', 'admin').where('active', '==', True).get()
            if len(admin_users) <= 1:
                return jsonify({'success': False, 'message': 'Cannot delete the last admin user'})
        
        # Delete user
        user_ref.delete()
        
        # Log activity
        activity_data = {
            'action': 'User Deleted',
            'details': f'Deleted user: {username}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'User deleted successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting user: {str(e)}'})

@app.route('/admin/hard_reset', methods=['POST'])
@admin_required
def hard_reset():
    """Hard reset - delete all data except admin users"""
    try:
        data = request.get_json()
        confirmation = data.get('confirmation', '')
        
        # Require explicit confirmation
        if confirmation != 'RESET ALL DATA':
            return jsonify({'success': False, 'message': 'Invalid confirmation. Type "RESET ALL DATA" to confirm.'})
        
        # Get current admin user info before reset
        current_user_id = session.get('user_id')
        current_username = session.get('username')
        
        # Collections to reset (keep users for admin access)
        collections_to_reset = [
            'products',
            'categories', 
            'customers',
            'sales_orders',
            'production_orders',
            'activities'
        ]
        
        deleted_counts = {}
        
        # Delete all documents from each collection
        for collection_name in collections_to_reset:
            try:
                collection_ref = db.collection(collection_name)
                docs = collection_ref.get()
                count = 0
                
                for doc in docs:
                    doc.reference.delete()
                    count += 1
                
                deleted_counts[collection_name] = count
                print(f"Deleted {count} documents from {collection_name}")
                
            except Exception as e:
                print(f"Error deleting from {collection_name}: {str(e)}")
                deleted_counts[collection_name] = f"Error: {str(e)}"
        
        # Clear all caches
        for cache_key in cache:
            cache[cache_key]['data'] = None
            cache[cache_key]['timestamp'] = 0
        
        # Create reset activity log
        try:
            activity_data = {
                'action': 'HARD RESET',
                'details': f'System reset by {current_username}. Deleted: {deleted_counts}',
                'user': current_username,
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
        except:
            pass  # If activities collection was deleted, skip logging
        
        # Log to console for debugging
        print(f"HARD RESET completed by {current_username}")
        print(f"Deleted counts: {deleted_counts}")
        
        return jsonify({
            'success': True, 
            'message': 'Hard reset completed successfully!',
            'deleted_counts': deleted_counts
        })
        
    except Exception as e:
        print(f"Error during hard reset: {str(e)}")
        return jsonify({'success': False, 'message': f'Error during reset: {str(e)}'})

@app.route('/excel_import', methods=['GET', 'POST'])
@login_required
def excel_import():
    if request.method == 'POST':
        try:
            file = request.files['file']
            if file and file.filename.endswith('.xlsx'):
                # Read Excel file
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                imported_count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Check if first column has data
                        product_data = {
                            'name': row[0],
                            'category': row[1] or '',
                            'size': row[2] or '',
                            'color': row[3] or '',
                            'price': float(row[4]) if row[4] else 0.0,
                            'body_size': str(row[5]) if row[5] else '',
                            'waist_size': str(row[6]) if len(row) > 6 and row[6] else '',
                            'length': str(row[7]) if len(row) > 7 and row[7] else '',
                            'description': row[8] if len(row) > 8 and row[8] else '',
                            'created_at': datetime.now(),
                            'updated_at': datetime.now()
                        }
                        
                        # Generate barcode
                        barcode_data = f"PROD_{uuid.uuid4().hex[:12].upper()}"
                        product_data['barcode'] = barcode_data
                        
                        db.collection('products').add(product_data)
                        imported_count += 1
                
                # Log activity
                activity_data = {
                    'action': 'Excel Import',
                    'details': f'Imported {imported_count} products from Excel',
                    'user': session['username'],
                    'timestamp': datetime.now()
                }
                db.collection('activities').add(activity_data)
                
                flash(f'Successfully imported {imported_count} products!', 'success')
            else:
                flash('Please upload a valid Excel file (.xlsx)', 'error')
                
        except Exception as e:
            flash(f'Error importing file: {str(e)}', 'error')
    
    return render_template('excel_import.html')

@app.route('/excel_export')
@login_required
def excel_export():
    try:
        # Get all products
        products_ref = db.collection('products')
        products = products_ref.get()
        
        # Create Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        
        # Add headers
        headers = ['Name', 'Category', 'Size', 'Color', 'Price', 'Body Size', 'Waist Size', 'Length', 'Description', 'Barcode']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Add product data
        for row, product in enumerate(products, 2):
            product_data = product.to_dict()
            sheet.cell(row=row, column=1, value=product_data.get('name', ''))
            sheet.cell(row=row, column=2, value=product_data.get('category', ''))
            sheet.cell(row=row, column=3, value=product_data.get('size', ''))
            sheet.cell(row=row, column=4, value=product_data.get('color', ''))
            sheet.cell(row=row, column=5, value=product_data.get('price', 0))
            sheet.cell(row=row, column=6, value=product_data.get('body_size', ''))
            sheet.cell(row=row, column=7, value=product_data.get('waist_size', ''))
            sheet.cell(row=row, column=8, value=product_data.get('length', ''))
            sheet.cell(row=row, column=9, value=product_data.get('description', ''))
            sheet.cell(row=row, column=10, value=product_data.get('barcode', ''))
        
        # Save to memory
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Log activity
        activity_data = {
            'action': 'Excel Export',
            'details': f'Exported {len(products)} products to Excel',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return output.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename=products_export.xlsx'
        }
        
    except Exception as e:
        flash(f'Error exporting file: {str(e)}', 'error')
        return redirect(url_for('products'))

# Production Management Routes
@app.route('/production')
@login_required
@performance_monitor
def production():
    # Get products from cache
    def fetch_products():
        products_ref = db.collection('products')
        products = products_ref.get()
        product_list = []
        
        for product in products:
            product_data = product.to_dict()
            product_data['id'] = product.id
            product_list.append(product_data)
        
        return product_list
    
    # Get production orders from cache
    def fetch_production():
        production_ref = db.collection('production_orders')
        production_orders = production_ref.get()
        production_list = []
        
        for order in production_orders:
            order_data = order.to_dict()
            order_data['id'] = order.id
            production_list.append(order_data)
        
        return production_list
    
    product_list = get_cached_data('products', fetch_products)
    production_list = get_cached_data('production_orders', fetch_production)
    
    return render_template('production.html', products=product_list, production_orders=production_list)

@app.route('/create_production', methods=['POST'])
@login_required
def create_production():
    try:
        notes = request.form.get('notes', '')
        created_orders = []
        
        # Process multiple products
        for key, value in request.form.items():
            if key.startswith('product_') and key.endswith('_quantity') and value:
                product_id = key.replace('product_', '').replace('_quantity', '')
                quantity = int(value)
                
                if quantity > 0:
                    # Get product details
                    product_ref = db.collection('products').document(product_id)
                    product_doc = product_ref.get()
                    
                    if product_doc.exists:
                        product_data = product_doc.to_dict()
                        
                        # Create production order
                        production_data = {
                            'product_id': product_id,
                            'product_name': product_data['name'],
                            'product_category': product_data['category'],
                            'product_size': product_data['size'],
                            'product_color': product_data['color'],
                            'status': 'pending',
                            'notes': notes,
                            'created_by': session['username'],
                            'created_at': datetime.now(),
                            'updated_at': datetime.now(),
                            'order_type': 'multiple' if len([k for k in request.form.keys() if k.startswith('product_') and k.endswith('_quantity') and request.form[k]]) > 1 else 'single'
                        }
                        
                        db.collection('production_orders').add(production_data)
                        created_orders.append(f"{product_data['name']} (Qty: {quantity})")
        
        if created_orders:
            # Invalidate cache
            cache['production_orders']['data'] = None
            
            # Log activity
            activity_data = {
                'action': 'Production Order Created',
                'details': f'Created production order for: {", ".join(created_orders)}',
                'user': session['username'],
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
            
            flash(f'Production order created successfully for {len(created_orders)} product(s)!', 'success')
        else:
            flash('No products selected or invalid quantities', 'error')
            
    except Exception as e:
        flash(f'Error creating production order: {str(e)}', 'error')
    
    return redirect(url_for('production'))

@app.route('/update_production_status/<order_id>', methods=['POST'])
@login_required
def update_production_status(order_id):
    try:
        status = request.form['status']
        
        production_ref = db.collection('production_orders').document(order_id)
        production_doc = production_ref.get()
        
        if production_doc.exists:
            production_data = production_doc.to_dict()
            
            # Update production order
            production_ref.update({
                'status': status,
                'updated_at': datetime.now(),
                'updated_by': session['username']
            })
            
            # Log activity
            activity_data = {
                'action': 'Production Status Updated',
                'details': f'Updated production order status to: {status}',
                'user': session['username'],
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
            
            flash('Production status updated successfully!', 'success')
        else:
            flash('Production order not found', 'error')
            
    except Exception as e:
        flash(f'Error updating production status: {str(e)}', 'error')
    
    return redirect(url_for('production'))

# Sales Management Routes
@app.route('/sales')
@login_required
@permission_required('sales_customer')
@performance_monitor
def sales():
    # Get products from cache
    def fetch_products():
        products_ref = db.collection('products')
        products = products_ref.get()
        product_list = []
        
        for product in products:
            product_data = product.to_dict()
            product_data['id'] = product.id
            product_list.append(product_data)
        
        return product_list
    
    # Get sales orders from cache
    def fetch_sales():
        sales_ref = db.collection('sales_orders')
        sales_orders = sales_ref.get()
        sales_list = []
        
        for order in sales_orders:
            order_data = order.to_dict()
            order_data['id'] = order.id
            sales_list.append(order_data)
        
        return sales_list
    
    # Group sales by customer with enhanced consolidation
    def group_sales_by_customer(sales_list):
        grouped_sales = {}
        for sale in sales_list:
            # Check if this is a new consolidated multiple items order
            if sale.get('is_multiple_items'):
                # This is already a consolidated order, treat as single group
                customer_key = f"{sale['customer_name']}|{sale['customer_phone']}|{sale.get('order_id', sale.get('created_at'))}"
                grouped_sales[customer_key] = {
                    'customer_name': sale['customer_name'],
                    'customer_phone': sale['customer_phone'],
                    'sold_by': sale['sold_by'],
                    'created_at': sale['created_at'],
                    'items': [sale]  # Single consolidated item
                }
            else:
                # Legacy individual items - group by customer and time (within 5 minutes)
                customer_base_key = f"{sale['customer_name']}|{sale['customer_phone']}"
                
                # Find existing group within 5 minutes
                found_group = False
                for existing_key in grouped_sales:
                    if existing_key.startswith(customer_base_key):
                        existing_group = grouped_sales[existing_key]
                        # Check if within 5 minutes of each other
                        time_diff = abs((sale['created_at'] - existing_group['created_at']).total_seconds())
                        if time_diff <= 300:  # 5 minutes
                            existing_group['items'].append(sale)
                            found_group = True
                            break
                
                if not found_group:
                    # Create new group for this customer/time
                    customer_key = f"{customer_base_key}|{sale['created_at']}"
                    grouped_sales[customer_key] = {
                        'customer_name': sale['customer_name'],
                        'customer_phone': sale['customer_phone'],
                        'sold_by': sale['sold_by'],
                        'created_at': sale['created_at'],
                        'items': [sale]
                    }
        
        # Convert to list and sort by date (newest first)
        grouped_list = list(grouped_sales.values())
        grouped_list.sort(key=lambda x: x['created_at'], reverse=True)
        return grouped_list
    
    # Get sizes and colors from cache
    def fetch_sizes():
        sizes_ref = db.collection('sizes')
        sizes = sizes_ref.get()
        size_list = []
        
        for size in sizes:
            size_data = size.to_dict()
            size_data['id'] = size.id
            size_list.append(size_data)
        
        return size_list
    
    def fetch_colors():
        colors_ref = db.collection('colors')
        colors = colors_ref.get()
        color_list = []
        
        for color in colors:
            color_data = color.to_dict()
            color_data['id'] = color.id
            color_list.append(color_data)
        
        return color_list
    
    product_list = get_cached_data('products', fetch_products)
    sales_list = get_cached_data('sales_orders', fetch_sales)
    grouped_sales = group_sales_by_customer(sales_list)
    size_list = get_cached_data('sizes', fetch_sizes)
    color_list = get_cached_data('colors', fetch_colors)
    
    return render_template('sales.html', products=product_list, sales_orders=sales_list, grouped_sales=grouped_sales, sizes=size_list, colors=color_list)

@app.route('/create_sale', methods=['POST'])
@login_required
@permission_required('sales_customer')
def create_sale():
    try:
        customer_name = request.form['customer_name']
        customer_address = request.form['customer_address']
        customer_phone = request.form['customer_phone']
        product_id = request.form['product_id']
        item_numbers = request.form['item_numbers']
        delivery_charge = float(request.form.get('delivery_charge', 0))
        emergency_delivery = request.form.get('emergency_delivery', 'false').lower() == 'true'
        notes = request.form.get('notes', '')
        sale_size = request.form.get('sale_size', '')
        sale_color = request.form.get('sale_color', '')
        
        # Parse item numbers to get quantity
        def parse_item_numbers(item_numbers_str):
            if not item_numbers_str:
                return 0
            
            total_items = 0
            parts = item_numbers_str.split(',')
            
            for part in parts:
                part = part.strip()
                if '-' in part:
                    # Handle ranges like "1-5"
                    try:
                        start, end = map(int, part.split('-'))
                        total_items += end - start + 1
                    except ValueError:
                        continue
                else:
                    # Handle single numbers
                    try:
                        total_items += 1
                    except ValueError:
                        continue
            
            return total_items
        
        quantity = parse_item_numbers(item_numbers)
        
        if quantity == 0:
            flash('Please enter valid item numbers', 'error')
            return redirect(url_for('sales'))
        
        # Get product details
        product_ref = db.collection('products').document(product_id)
        product_doc = product_ref.get()
        
        if product_doc.exists:
            product_data = product_doc.to_dict()
            
            # Calculate total price
            product_total = product_data['price'] * quantity
            total_price = product_total + delivery_charge
            
            # Use override values if provided, otherwise use product defaults
            final_size = sale_size if sale_size else product_data.get('size', '')
            final_color = sale_color if sale_color else product_data['color']
            
            # Create sales order
            sale_data = {
                'customer_name': customer_name,
                'customer_address': customer_address,
                'customer_phone': customer_phone,
                'product_id': product_id,
                'product_name': product_data['name'],
                'product_category': product_data['category'],
                'product_size': final_size,
                'product_body_size': product_data.get('body_size', ''),
                'product_waist_size': product_data.get('waist_size', ''),
                'product_length': product_data.get('length', ''),
                'product_color': final_color,
                'product_price': product_data['price'],
                'quantity': quantity,
                'item_numbers': item_numbers,
                'product_total': product_total,
                'delivery_charge': delivery_charge,
                'total_price': total_price,
                'status': 'completed',
                'emergency_delivery': emergency_delivery,
                'notes': notes,
                'sold_by': session['username'],
                'created_at': datetime.now()
            }
            
            db.collection('sales_orders').add(sale_data)
            
            # Invalidate caches
            cache['sales_orders']['data'] = None
            cache['products']['data'] = None
            
            
            # Log activity
            activity_data = {
                'action': 'Sale Completed',
                'details': f'Sold {quantity} units of {product_data["name"]} to {customer_name} - Items: {item_numbers} - Total: ৳{total_price}',
                'user': session['username'],
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
            
            flash(f'Sale completed successfully! Items: {item_numbers} - Total: ৳{total_price}', 'success')
        else:
            flash('Product not found', 'error')
            
    except Exception as e:
        flash(f'Error creating sale: {str(e)}', 'error')
    
    return redirect(url_for('sales'))

# Get Sale Details for Receipt
@app.route('/get_sale_details/<sale_id>')
@login_required
@permission_required('sales_customer')
def get_sale_details(sale_id):
    try:
        sale_ref = db.collection('sales_orders').document(sale_id)
        sale_doc = sale_ref.get()
        
        if sale_doc.exists:
            sale_data = sale_doc.to_dict()
            sale_data['id'] = sale_doc.id
            return jsonify({'success': True, 'sale': sale_data})
        else:
            return jsonify({'success': False, 'message': 'Sale not found'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Customer Management Routes
@app.route('/customers')
@login_required
@permission_required('sales_customer')
def customers():
    customers_ref = db.collection('customers')
    customers = customers_ref.get()
    customer_list = []
    
    for customer in customers:
        customer_data = customer.to_dict()
        customer_data['id'] = customer.id
        customer_list.append(customer_data)
    
    return render_template('customers.html', customers=customer_list)

@app.route('/add_customer', methods=['POST'])
@login_required
@permission_required('sales_customer')
def add_customer():
    try:
        name = request.form['name']
        address = request.form['address']
        phone = request.form['phone']
        email = request.form.get('email', '')
        
        customer_data = {
            'name': name,
            'address': address,
            'phone': phone,
            'email': email,
            'created_at': datetime.now(),
            'created_by': session['username']
        }
        
        db.collection('customers').add(customer_data)
        
        # Log activity
        activity_data = {
            'action': 'Customer Added',
            'details': f'Added customer: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        flash('Customer added successfully!', 'success')
        
    except Exception as e:
        flash(f'Error adding customer: {str(e)}', 'error')
    
    return redirect(url_for('customers'))

# Category Management Routes
@app.route('/categories')
@login_required
@performance_monitor
def categories():
    def fetch_categories():
        categories_ref = db.collection('categories')
        categories = categories_ref.get()
        category_list = []
        
        for category in categories:
            category_data = category.to_dict()
            category_data['id'] = category.id
            category_list.append(category_data)
        
        return category_list
    
    category_list = get_cached_data('categories', fetch_categories)
    return render_template('categories.html', categories=category_list)

@app.route('/add_category', methods=['POST'])
@login_required
def add_category():
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        # Check if category already exists
        categories_ref = db.collection('categories')
        existing = categories_ref.where('name', '==', name).get()
        
        if existing:
            flash('Category already exists!', 'error')
            return redirect(url_for('categories'))
        
        category_data = {
            'name': name,
            'description': description,
            'created_at': datetime.now(),
            'created_by': session['username']
        }
        
        db.collection('categories').add(category_data)
        
        # Invalidate cache
        cache['categories']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Category Added',
            'details': f'Added category: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        flash('Category added successfully!', 'success')
        
    except Exception as e:
        flash(f'Error adding category: {str(e)}', 'error')
    
    return redirect(url_for('categories'))

@app.route('/add_category_ajax', methods=['POST'])
@login_required
def add_category_ajax():
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        # Check if category already exists
        categories_ref = db.collection('categories')
        existing = categories_ref.where('name', '==', name).get()
        
        if existing:
            return jsonify({'success': False, 'message': 'Category already exists!'})
        
        category_data = {
            'name': name,
            'description': description,
            'created_at': datetime.now(),
            'created_by': session['username']
        }
        
        db.collection('categories').add(category_data)
        
        # Invalidate cache
        cache['categories']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Category Added',
            'details': f'Added category: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Category added successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error adding category: {str(e)}'})

@app.route('/edit_category_ajax/<category_id>', methods=['POST'])
@login_required
def edit_category_ajax(category_id):
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        category_ref = db.collection('categories').document(category_id)
        category_doc = category_ref.get()
        
        if not category_doc.exists:
            return jsonify({'success': False, 'message': 'Category not found!'})
        
        # Check if new name conflicts with existing category
        if name != category_doc.to_dict()['name']:
            categories_ref = db.collection('categories')
            existing = categories_ref.where('name', '==', name).get()
            if existing:
                return jsonify({'success': False, 'message': 'Category name already exists!'})
        
        # Update category
        category_ref.update({
            'name': name,
            'description': description,
            'updated_at': datetime.now(),
            'updated_by': session['username']
        })
        
        # Invalidate cache
        cache['categories']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Category Updated',
            'details': f'Updated category: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Category updated successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error updating category: {str(e)}'})

# Size Management Routes
@app.route('/sizes')
@login_required
@permission_required('manage_products')
def sizes():
    # Get sizes from cache
    def fetch_sizes():
        sizes_ref = db.collection('sizes')
        sizes = sizes_ref.get()
        size_list = []
        
        for size in sizes:
            size_data = size.to_dict()
            size_data['id'] = size.id
            size_list.append(size_data)
        
        return size_list
    
    size_list = get_cached_data('sizes', fetch_sizes)
    return render_template('sizes.html', sizes=size_list)

@app.route('/add_size', methods=['POST'])
@login_required
@permission_required('manage_products')
def add_size():
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        # Check if size already exists
        sizes_ref = db.collection('sizes')
        existing = sizes_ref.where('name', '==', name).get()
        
        if existing:
            flash('Size already exists!', 'error')
            return redirect(url_for('sizes'))
        
        size_data = {
            'name': name,
            'description': description,
            'created_at': datetime.now(),
            'created_by': session['username']
        }
        
        db.collection('sizes').add(size_data)
        
        # Invalidate cache
        cache['sizes']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Size Added',
            'details': f'Added size: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        flash('Size added successfully!', 'success')
        
    except Exception as e:
        flash(f'Error adding size: {str(e)}', 'error')
    
    return redirect(url_for('sizes'))

@app.route('/add_size_ajax', methods=['POST'])
@login_required
@permission_required('sales_customer')
def add_size_ajax():
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        # Check if size already exists
        sizes_ref = db.collection('sizes')
        existing = sizes_ref.where('name', '==', name).get()
        
        if existing:
            return jsonify({'success': False, 'message': 'Size already exists!'})
        
        size_data = {
            'name': name,
            'description': description,
            'created_at': datetime.now(),
            'created_by': session['username']
        }
        
        db.collection('sizes').add(size_data)
        
        # Invalidate cache
        cache['sizes']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Size Added',
            'details': f'Added size: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Size added successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error adding size: {str(e)}'})

@app.route('/get_sizes_ajax')
@login_required
@permission_required('sales_customer')
def get_sizes_ajax():
    try:
        def fetch_sizes():
            sizes_ref = db.collection('sizes')
            sizes = sizes_ref.get()
            size_list = []
            
            for size in sizes:
                size_data = size.to_dict()
                size_data['id'] = size.id
                size_list.append(size_data)
            
            return size_list
        
        size_list = get_cached_data('sizes', fetch_sizes)
        return jsonify({'success': True, 'sizes': size_list})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error fetching sizes: {str(e)}'})

@app.route('/edit_size_ajax/<size_id>', methods=['POST'])
@login_required
@permission_required('sales_customer')
def edit_size_ajax(size_id):
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        size_ref = db.collection('sizes').document(size_id)
        size_doc = size_ref.get()
        
        if not size_doc.exists:
            return jsonify({'success': False, 'message': 'Size not found!'})
        
        # Check if new name conflicts with existing size
        if name != size_doc.to_dict()['name']:
            sizes_ref = db.collection('sizes')
            existing = sizes_ref.where('name', '==', name).get()
            if existing:
                return jsonify({'success': False, 'message': 'Size name already exists!'})
        
        # Update size
        size_ref.update({
            'name': name,
            'description': description,
            'updated_at': datetime.now(),
            'updated_by': session['username']
        })
        
        # Invalidate cache
        cache['sizes']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Size Updated',
            'details': f'Updated size: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Size updated successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error updating size: {str(e)}'})

@app.route('/delete_size_ajax/<size_id>', methods=['POST'])
@login_required
@permission_required('sales_customer')
def delete_size_ajax(size_id):
    try:
        size_ref = db.collection('sizes').document(size_id)
        size_doc = size_ref.get()
        
        if not size_doc.exists:
            return jsonify({'success': False, 'message': 'Size not found!'})
        
        size_data = size_doc.to_dict()
        size_name = size_data['name']
        
        # Check if size is being used in products
        products_ref = db.collection('products')
        products_using_size = products_ref.where('size', '==', size_name).get()
        
        if products_using_size:
            return jsonify({'success': False, 'message': f'Cannot delete size "{size_name}" as it is being used in {len(products_using_size)} product(s)!'})
        
        # Delete size
        size_ref.delete()
        
        # Invalidate cache
        cache['sizes']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Size Deleted',
            'details': f'Deleted size: {size_name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Size deleted successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting size: {str(e)}'})

# Color Management Routes
@app.route('/colors')
@login_required
@permission_required('manage_products')
def colors():
    # Get colors from cache
    def fetch_colors():
        colors_ref = db.collection('colors')
        colors = colors_ref.get()
        color_list = []
        
        for color in colors:
            color_data = color.to_dict()
            color_data['id'] = color.id
            color_list.append(color_data)
        
        return color_list
    
    color_list = get_cached_data('colors', fetch_colors)
    return render_template('colors.html', colors=color_list)

@app.route('/add_color', methods=['POST'])
@login_required
@permission_required('manage_products')
def add_color():
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        # Check if color already exists
        colors_ref = db.collection('colors')
        existing = colors_ref.where('name', '==', name).get()
        
        if existing:
            flash('Color already exists!', 'error')
            return redirect(url_for('colors'))
        
        color_data = {
            'name': name,
            'description': description,
            'created_at': datetime.now(),
            'created_by': session['username']
        }
        
        db.collection('colors').add(color_data)
        
        # Invalidate cache
        cache['colors']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Color Added',
            'details': f'Added color: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        flash('Color added successfully!', 'success')
        
    except Exception as e:
        flash(f'Error adding color: {str(e)}', 'error')
    
    return redirect(url_for('colors'))

@app.route('/add_color_ajax', methods=['POST'])
@login_required
@permission_required('sales_customer')
def add_color_ajax():
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        # Check if color already exists
        colors_ref = db.collection('colors')
        existing = colors_ref.where('name', '==', name).get()
        
        if existing:
            return jsonify({'success': False, 'message': 'Color already exists!'})
        
        color_data = {
            'name': name,
            'description': description,
            'created_at': datetime.now(),
            'created_by': session['username']
        }
        
        db.collection('colors').add(color_data)
        
        # Invalidate cache
        cache['colors']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Color Added',
            'details': f'Added color: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Color added successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error adding color: {str(e)}'})

@app.route('/get_colors_ajax')
@login_required
@permission_required('sales_customer')
def get_colors_ajax():
    try:
        def fetch_colors():
            colors_ref = db.collection('colors')
            colors = colors_ref.get()
            color_list = []
            
            for color in colors:
                color_data = color.to_dict()
                color_data['id'] = color.id
                color_list.append(color_data)
            
            return color_list
        
        color_list = get_cached_data('colors', fetch_colors)
        return jsonify({'success': True, 'colors': color_list})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error fetching colors: {str(e)}'})

@app.route('/edit_color_ajax/<color_id>', methods=['POST'])
@login_required
@permission_required('sales_customer')
def edit_color_ajax(color_id):
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        color_ref = db.collection('colors').document(color_id)
        color_doc = color_ref.get()
        
        if not color_doc.exists:
            return jsonify({'success': False, 'message': 'Color not found!'})
        
        # Check if new name conflicts with existing color
        if name != color_doc.to_dict()['name']:
            colors_ref = db.collection('colors')
            existing = colors_ref.where('name', '==', name).get()
            if existing:
                return jsonify({'success': False, 'message': 'Color name already exists!'})
        
        # Update color
        color_ref.update({
            'name': name,
            'description': description,
            'updated_at': datetime.now(),
            'updated_by': session['username']
        })
        
        # Invalidate cache
        cache['colors']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Color Updated',
            'details': f'Updated color: {name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Color updated successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error updating color: {str(e)}'})

@app.route('/delete_color_ajax/<color_id>', methods=['POST'])
@login_required
@permission_required('sales_customer')
def delete_color_ajax(color_id):
    try:
        color_ref = db.collection('colors').document(color_id)
        color_doc = color_ref.get()
        
        if not color_doc.exists:
            return jsonify({'success': False, 'message': 'Color not found!'})
        
        color_data = color_doc.to_dict()
        color_name = color_data['name']
        
        # Check if color is being used in products
        products_ref = db.collection('products')
        products_using_color = products_ref.where('color', '==', color_name).get()
        
        if products_using_color:
            return jsonify({'success': False, 'message': f'Cannot delete color "{color_name}" as it is being used in {len(products_using_color)} product(s)!'})
        
        # Delete color
        color_ref.delete()
        
        # Invalidate cache
        cache['colors']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Color Deleted',
            'details': f'Deleted color: {color_name}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Color deleted successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting color: {str(e)}'})

@app.route('/delete_category_ajax/<category_id>', methods=['POST'])
@login_required
def delete_category_ajax(category_id):
    try:
        category_ref = db.collection('categories').document(category_id)
        category_doc = category_ref.get()
        
        if not category_doc.exists:
            return jsonify({'success': False, 'message': 'Category not found!'})
        
        category_data = category_doc.to_dict()
        
        # Check if category is used in products
        products_ref = db.collection('products')
        products_using_category = products_ref.where('category', '==', category_data['name']).get()
        
        if products_using_category:
            return jsonify({
                'success': False, 
                'message': f'Cannot delete category "{category_data["name"]}" - it is used by {len(products_using_category)} products'
            })
        
        # Delete category
        category_ref.delete()
        
        # Invalidate cache
        cache['categories']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Category Deleted',
            'details': f'Deleted category: {category_data["name"]}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({'success': True, 'message': 'Category deleted successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error deleting category: {str(e)}'})

@app.route('/edit_category/<category_id>', methods=['POST'])
@login_required
def edit_category(category_id):
    try:
        name = request.form['name']
        description = request.form.get('description', '')
        
        category_ref = db.collection('categories').document(category_id)
        category_doc = category_ref.get()
        
        if category_doc.exists:
            category_ref.update({
                'name': name,
                'description': description,
                'updated_at': datetime.now(),
                'updated_by': session['username']
            })
            
            # Log activity
            activity_data = {
                'action': 'Category Updated',
                'details': f'Updated category: {name}',
                'user': session['username'],
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
            
            flash('Category updated successfully!', 'success')
        else:
            flash('Category not found', 'error')
            
    except Exception as e:
        flash(f'Error updating category: {str(e)}', 'error')
    
    return redirect(url_for('categories'))

@app.route('/delete_category/<category_id>')
@login_required
def delete_category(category_id):
    try:
        category_ref = db.collection('categories').document(category_id)
        category_doc = category_ref.get()
        
        if category_doc.exists:
            category_data = category_doc.to_dict()
            
            # Check if category is used in products
            products_ref = db.collection('products')
            products_using_category = products_ref.where('category', '==', category_data['name']).get()
            
            if products_using_category:
                flash(f'Cannot delete category "{category_data["name"]}" - it is used by {len(products_using_category)} products', 'error')
                return redirect(url_for('categories'))
            
            category_ref.delete()
            
            # Log activity
            activity_data = {
                'action': 'Category Deleted',
                'details': f'Deleted category: {category_data["name"]}',
                'user': session['username'],
                'timestamp': datetime.now()
            }
            db.collection('activities').add(activity_data)
            
            flash('Category deleted successfully!', 'success')
        else:
            flash('Category not found', 'error')
            
    except Exception as e:
        flash(f'Error deleting category: {str(e)}', 'error')
    
    return redirect(url_for('categories'))

# Excel Import for Production
@app.route('/excel_import_production', methods=['GET', 'POST'])
@login_required
def excel_import_production():
    if request.method == 'POST':
        try:
            file = request.files['file']
            if file and file.filename.endswith('.xlsx'):
                # Read Excel file
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                imported_count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # Check if first column has data
                        # Find product by name
                        products_ref = db.collection('products')
                        product_query = products_ref.where('name', '==', row[0]).get()
                        
                        if product_query:
                            product = product_query[0]
                            product_data = product.to_dict()
                            
                            production_data = {
                                'product_id': product.id,
                                'product_name': product_data['name'],
                                'product_category': product_data['category'],
                                'product_size': product_data['size'],
                                'product_color': product_data['color'],
                                'quantity': int(row[1]) if row[1] else 0,
                                'status': row[2] if row[2] else 'pending',
                                'notes': row[3] if row[3] else '',
                                'created_by': session['username'],
                                'created_at': datetime.now(),
                                'updated_at': datetime.now()
                            }
                            
                            db.collection('production_orders').add(production_data)
                            imported_count += 1
                
                # Log activity
                activity_data = {
                    'action': 'Excel Production Import',
                    'details': f'Imported {imported_count} production orders from Excel',
                    'user': session['username'],
                    'timestamp': datetime.now()
                }
                db.collection('activities').add(activity_data)
                
                flash(f'Successfully imported {imported_count} production orders!', 'success')
            else:
                flash('Please upload a valid Excel file (.xlsx)', 'error')
                
        except Exception as e:
            flash(f'Error importing file: {str(e)}', 'error')
    
    return render_template('excel_import_production.html')

# Excel Export for Production Records
@app.route('/excel_export_production')
@login_required
def excel_export_production():
    try:
        # Get date filter
        date_filter = request.args.get('date', '')
        
        # Get production orders
        production_ref = db.collection('production_orders')
        production_orders = production_ref.get()
        
        # Filter by date if provided
        filtered_orders = []
        for order in production_orders:
            order_data = order.to_dict()
            if date_filter:
                order_date = order_data.get('created_at', datetime.now())
                if isinstance(order_date, datetime):
                    if order_date.strftime('%Y-%m-%d') != date_filter:
                        continue
            order_data['id'] = order.id
            filtered_orders.append(order_data)
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Production Records"
        
        # Headers
        headers = ['Product Name', 'Category', 'Size', 'Color', 'Quantity', 'Status', 'Notes', 'Created By', 'Created Date', 'Updated Date']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, order in enumerate(filtered_orders, 2):
            sheet.cell(row=row, column=1, value=order.get('product_name', ''))
            sheet.cell(row=row, column=2, value=order.get('product_category', ''))
            sheet.cell(row=row, column=3, value=order.get('product_size', ''))
            sheet.cell(row=row, column=4, value=order.get('product_color', ''))
            sheet.cell(row=row, column=5, value=order.get('quantity', 0))
            sheet.cell(row=row, column=6, value=order.get('status', ''))
            sheet.cell(row=row, column=7, value=order.get('notes', ''))
            sheet.cell(row=row, column=8, value=order.get('created_by', ''))
            sheet.cell(row=row, column=9, value=order.get('created_at', '').strftime('%Y-%m-%d %H:%M') if order.get('created_at') else '')
            sheet.cell(row=row, column=10, value=order.get('updated_at', '').strftime('%Y-%m-%d %H:%M') if order.get('updated_at') else '')
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'production_records_{date_filter if date_filter else "all"}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting production records: {str(e)}', 'error')
        return redirect(url_for('production'))

# Excel Export for Sales (Delivery)
@app.route('/excel_export_sales')
@login_required
@permission_required('sales_customer')
def excel_export_sales():
    try:
        # Get date filter
        date_filter = request.args.get('date', '')
        
        # Get sales orders
        sales_ref = db.collection('sales_orders')
        sales_orders = sales_ref.get()
        
        # Filter by date if provided
        filtered_orders = []
        for order in sales_orders:
            order_data = order.to_dict()
            if date_filter:
                order_date = order_data.get('created_at', datetime.now())
                if isinstance(order_date, datetime):
                    if order_date.strftime('%Y-%m-%d') != date_filter:
                        continue
            filtered_orders.append(order_data)
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sales Delivery"
        
        # Headers - Consolidated format
        headers = ['Customer Name', 'Phone', 'Address', 'All Products', 'Total Items', 'Total Amount', 'Sale Date', 'Sold By']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Group sales by customer and consolidate all their items
        grouped_sales = {}
        for order_data in filtered_orders:
            customer_key = f"{order_data.get('customer_name', '')}|{order_data.get('customer_phone', '')}"
            if customer_key not in grouped_sales:
                grouped_sales[customer_key] = {
                    'customer_name': order_data.get('customer_name', ''),
                    'customer_phone': order_data.get('customer_phone', ''),
                    'customer_address': order_data.get('customer_address', ''),
                    'sold_by': order_data.get('sold_by', ''),
                    'created_at': order_data.get('created_at', ''),
                    'items': [],
                    'total_amount': 0,
                    'total_items': 0
                }
            
            # Handle both consolidated multiple items and individual items
            if order_data.get('is_multiple_items'):
                # This is a consolidated multiple items order
                grouped_sales[customer_key]['items'].extend(order_data.get('items', []))
                grouped_sales[customer_key]['total_amount'] += order_data.get('total_price', 0)
                grouped_sales[customer_key]['total_items'] += order_data.get('total_quantity', 0)
            else:
                # This is a single item order
                grouped_sales[customer_key]['items'].append(order_data)
                grouped_sales[customer_key]['total_amount'] += order_data.get('total_price', 0)
                grouped_sales[customer_key]['total_items'] += max(order_data.get('quantity', 0), 1)
        
        # Sort by date (newest first)
        grouped_list = list(grouped_sales.values())
        grouped_list.sort(key=lambda x: x['created_at'], reverse=True)
        
        # Data - One row per customer with all their products
        row = 2
        for customer_group in grouped_list:
            # Build consolidated product list
            product_list = []
            for item in customer_group['items']:
                if isinstance(item, dict):
                    # Handle both consolidated items and regular items
                    product_name = item.get('product_name', '')
                    product_size = item.get('product_size', '')
                    product_color = item.get('product_color', '')
                    quantity = item.get('quantity', item.get('item_numbers', 1))
                    
                    # Format: "Product Name (Size, Color) x Quantity"
                    product_desc = f"{product_name}"
                    if product_size or product_color:
                        details = []
                        if product_size:
                            details.append(product_size)
                        if product_color:
                            details.append(product_color)
                        product_desc += f" ({', '.join(details)})"
                    product_desc += f" x{quantity}"
                    product_list.append(product_desc)
            
            # Customer information
            sheet.cell(row=row, column=1, value=customer_group['customer_name'])
            sheet.cell(row=row, column=2, value=customer_group['customer_phone'])
            sheet.cell(row=row, column=3, value=customer_group['customer_address'])
            sheet.cell(row=row, column=4, value='; '.join(product_list))  # All products in one cell
            sheet.cell(row=row, column=5, value=customer_group['total_items'])
            sheet.cell(row=row, column=6, value=customer_group['total_amount'])
            sheet.cell(row=row, column=7, value=customer_group['created_at'].strftime('%Y-%m-%d %H:%M') if customer_group['created_at'] else '')
            sheet.cell(row=row, column=8, value=customer_group['sold_by'])
            
            row += 1
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'sales_delivery_{date_filter if date_filter else "all"}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting sales: {str(e)}', 'error')
        return redirect(url_for('sales'))

# Excel Export for Production
@app.route('/excel_export_to_production')
@login_required
@permission_required('sales_customer')
def excel_export_to_production():
    try:
        # Get date filter
        date_filter = request.args.get('date', '')
        
        # Get sales orders
        sales_ref = db.collection('sales_orders')
        sales_orders = sales_ref.get()
        
        # Filter by date if provided
        filtered_orders = []
        for order in sales_orders:
            order_data = order.to_dict()
            if date_filter:
                order_date = order_data.get('created_at', datetime.now())
                if isinstance(order_date, datetime):
                    if order_date.strftime('%Y-%m-%d') != date_filter:
                        continue
            filtered_orders.append(order_data)
        
        # Group products by their characteristics and sum quantities
        product_groups = {}
        
        for order_data in filtered_orders:
            # Handle both consolidated multiple items and individual items
            items_to_process = []
            
            if order_data.get('is_multiple_items'):
                # This is a consolidated multiple items order - process each item
                items_to_process = order_data.get('items', [])
            else:
                # This is a single item order
                items_to_process = [order_data]
            
            # Process each item
            for item in items_to_process:
                # Create a unique key for each product variant
                product_key = (
                    item.get('product_name', ''),
                    item.get('product_category', ''),
                    item.get('product_color', ''),
                    item.get('product_size', ''),
                    item.get('product_body_size', ''),
                    item.get('product_waist_size', ''),
                    item.get('product_length', '')
                )
                
                if product_key not in product_groups:
                    product_groups[product_key] = {
                        'product_name': item.get('product_name', ''),
                        'product_category': item.get('product_category', ''),
                        'product_color': item.get('product_color', ''),
                        'product_size': item.get('product_size', ''),
                        'product_body_size': item.get('product_body_size', ''),
                        'product_waist_size': item.get('product_waist_size', ''),
                        'product_length': item.get('product_length', ''),
                        'total_quantity': 0,
                        'item_numbers': []
                    }
                
                # Sum quantities
                quantity = max(item.get('quantity', 0), 1)
                product_groups[product_key]['total_quantity'] += quantity
                
                # Collect item numbers
                item_numbers = item.get('item_numbers', '')
                if item_numbers:
                    product_groups[product_key]['item_numbers'].append(item_numbers)
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Production Export"
        
        # Headers for production export
        headers = ['Product Name', 'Product Category', 'Product Color', 'Product Size', 
                   'Body Size', 'Waist Size', 'Length', 'Item Numbers', 'Total Quantity']
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Sort products by name, then by size, then by color for better organization
        sorted_products = sorted(product_groups.values(), 
                               key=lambda x: (x['product_name'], x['product_category'], 
                                            x['product_size'], x['product_color']))
        
        # Add data
        row = 2
        for product_data in sorted_products:
            sheet.cell(row=row, column=1, value=product_data['product_name'])
            sheet.cell(row=row, column=2, value=product_data['product_category'])
            sheet.cell(row=row, column=3, value=product_data['product_color'])
            sheet.cell(row=row, column=4, value=product_data['product_size'])
            sheet.cell(row=row, column=5, value=product_data['product_body_size'])
            sheet.cell(row=row, column=6, value=product_data['product_waist_size'])
            sheet.cell(row=row, column=7, value=product_data['product_length'])
            sheet.cell(row=row, column=8, value=', '.join(product_data['item_numbers']) if product_data['item_numbers'] else '')
            sheet.cell(row=row, column=9, value=product_data['total_quantity'])
            row += 1
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'production_export_{date_filter if date_filter else "all"}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting to production: {str(e)}', 'error')
        return redirect(url_for('sales'))

# Multiple Item Sales
@app.route('/create_multiple_sale', methods=['POST'])
@login_required
@permission_required('sales_customer')
def create_multiple_sale():
    try:
        customer_name = request.form['customer_name']
        customer_address = request.form['customer_address']
        customer_phone = request.form['customer_phone']
        delivery_charge = float(request.form.get('delivery_charge', 0))
        notes = request.form.get('notes', '')
        
        # Get selected products - handle both old and new format
        selected_products = []
        total_amount = 0
        
        # Check if using new variant format
        variant_keys = [key for key in request.form.keys() if key.startswith('variant_') and key.endswith('_product')]
        
        if variant_keys:
            # New variant-based format
            for key in variant_keys:
                variant_id = key.replace('variant_', '').replace('_product', '')
                product_id = request.form.get(f'variant_{variant_id}_product')
                item_numbers = request.form.get(f'variant_{variant_id}_items', '').strip()
                
                if product_id and item_numbers:
                    # Get product details
                    product_ref = db.collection('products').document(product_id)
                    product_doc = product_ref.get()
                    
                    if product_doc.exists:
                        product_data = product_doc.to_dict()
                        
                        # Get size and color overrides for this variant
                        size_override = request.form.get(f'variant_{variant_id}_size', '')
                        color_override = request.form.get(f'variant_{variant_id}_color', '')
                        
                        # Use override values if provided, otherwise use product defaults
                        final_size = size_override if size_override else product_data.get('size', '')
                        final_color = color_override if color_override else product_data['color']
                        
                        # Parse item numbers to get count
                        item_count = parse_item_numbers(item_numbers)
                        item_total = product_data['price'] * item_count
                        total_amount += item_total
                        
                        selected_products.append({
                            'product_id': product_id,
                            'product_name': product_data['name'],
                            'product_category': product_data['category'],
                            'product_size': final_size,
                            'product_body_size': product_data.get('body_size', ''),
                            'product_waist_size': product_data.get('waist_size', ''),
                            'product_length': product_data.get('length', ''),
                            'product_color': final_color,
                            'product_price': product_data['price'],
                            'item_numbers': item_numbers,
                            'quantity': item_count,
                            'item_total': item_total,
                            'variant_id': variant_id
                        })
        else:
            # Original format - for backward compatibility
            for key, value in request.form.items():
                if key.startswith('product_') and key.endswith('_items') and value:
                    product_id = key.replace('product_', '').replace('_items', '')
                    item_numbers = value.strip()
                    
                    if item_numbers:
                        # Get product details
                        product_ref = db.collection('products').document(product_id)
                        product_doc = product_ref.get()
                        
                        if product_doc.exists:
                            product_data = product_doc.to_dict()
                            
                            # Get size and color overrides for this product
                            size_override = request.form.get(f'product_{product_id}_size', '')
                            color_override = request.form.get(f'product_{product_id}_color', '')
                            
                            # Use override values if provided, otherwise use product defaults
                            final_size = size_override if size_override else product_data.get('size', '')
                            final_color = color_override if color_override else product_data['color']
                            
                            # Parse item numbers to get count
                            item_count = parse_item_numbers(item_numbers)
                            item_total = product_data['price'] * item_count
                            total_amount += item_total
                            
                            selected_products.append({
                                'product_id': product_id,
                                'product_name': product_data['name'],
                                'product_category': product_data['category'],
                                'product_size': final_size,
                                'product_body_size': product_data.get('body_size', ''),
                                'product_waist_size': product_data.get('waist_size', ''),
                                'product_length': product_data.get('length', ''),
                                'product_color': final_color,
                                'product_price': product_data['price'],
                                'item_numbers': item_numbers,
                                'quantity': item_count,
                                'item_total': item_total
                            })
        
        if not selected_products:
            flash('Please select at least one product', 'error')
            return redirect(url_for('sales'))
        
        # Create a single consolidated sales order for all items
        grand_total = total_amount + delivery_charge
        
        # Generate a unique order ID for this multiple item sale
        order_id = str(uuid.uuid4())
        
        sale_data = {
            'customer_name': customer_name,
            'customer_address': customer_address,
            'customer_phone': customer_phone,
            'order_id': order_id,  # Unique identifier for this multi-item order
            'items': selected_products,  # Store all items in a single record
            'product_total': total_amount,  # Total of all products
            'delivery_charge': delivery_charge,  # Full delivery charge (not split)
            'total_price': grand_total,  # Product total + delivery charge
            'total_items': len(selected_products),  # Number of different items
            'total_quantity': sum(item['quantity'] for item in selected_products),  # Total quantity
            'status': 'completed',
            'notes': notes,
            'sold_by': session['username'],
            'created_at': datetime.now(),
            'order_type': 'multiple_consolidated',  # New order type for consolidated orders
            'is_multiple_items': True  # Flag to identify multiple item orders
        }
        
        # Add the consolidated sale record
        db.collection('sales_orders').add(sale_data)
            
        
        # Invalidate cache
        cache['sales_orders']['data'] = None
        cache['products']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Consolidated Multiple Sale Completed',
            'details': f'Sold {len(selected_products)} items ({sum(item["quantity"] for item in selected_products)} total qty) to {customer_name} - Order ID: {order_id} - Total: ৳{grand_total}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        flash(f'Multiple sale completed successfully! {len(selected_products)} items - Total: ৳{grand_total}', 'success')
        
    except Exception as e:
        flash(f'Error creating multiple sale: {str(e)}', 'error')
    
    return redirect(url_for('sales'))

# Mark Sale as Returned
@app.route('/mark_sale_returned/<sale_id>', methods=['POST'])
@login_required
@permission_required('sales_customer')
def mark_sale_returned(sale_id):
    try:
        # Update the sale status to returned
        sale_ref = db.collection('sales_orders').document(sale_id)
        sale_ref.update({
            'status': 'returned',
            'returned_at': datetime.now(),
            'returned_by': session['username']
        })
        
        # Invalidate sales cache to reflect changes immediately
        cache['sales_orders']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Sale Marked as Returned',
            'details': f'Sale {sale_id} marked as returned',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        flash('Sale marked as returned successfully!', 'success')
        
    except Exception as e:
        flash(f'Error marking sale as returned: {str(e)}', 'error')
    
    return redirect(url_for('sales'))

# Toggle Emergency Delivery
@app.route('/toggle_emergency_delivery/<sale_id>', methods=['POST'])
@login_required
@permission_required('sales_customer')
def toggle_emergency_delivery(sale_id):
    try:
        # Get current sale data
        sale_ref = db.collection('sales_orders').document(sale_id)
        sale_doc = sale_ref.get()
        
        if not sale_doc.exists:
            return jsonify({'success': False, 'message': 'Sale not found'})
        
        sale_data = sale_doc.to_dict()
        current_emergency = sale_data.get('emergency_delivery', False)
        
        # Toggle emergency delivery status
        new_emergency_status = not current_emergency
        
        sale_ref.update({
            'emergency_delivery': new_emergency_status,
            'updated_at': datetime.now(),
            'updated_by': session['username']
        })
        
        # Invalidate sales cache to reflect changes immediately
        cache['sales_orders']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Emergency Delivery Toggled',
            'details': f'Sale {sale_id} emergency delivery {"enabled" if new_emergency_status else "disabled"}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({
            'success': True, 
            'message': f'Emergency delivery {"enabled" if new_emergency_status else "disabled"} successfully!',
            'emergency_delivery': new_emergency_status
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error toggling emergency delivery: {str(e)}'})

# Toggle Delivered Status
@app.route('/toggle_delivered/<sale_id>', methods=['POST'])
@login_required
@permission_required('sales_customer')
def toggle_delivered(sale_id):
    try:
        # Get current sale data
        sale_ref = db.collection('sales_orders').document(sale_id)
        sale_doc = sale_ref.get()
        
        if not sale_doc.exists:
            return jsonify({'success': False, 'message': 'Sale not found'})
        
        sale_data = sale_doc.to_dict()
        current_delivered = sale_data.get('delivered', False)
        
        # Toggle delivered status
        new_delivered_status = not current_delivered
        
        sale_ref.update({
            'delivered': new_delivered_status,
            'delivered_at': datetime.now() if new_delivered_status else None,
            'updated_at': datetime.now(),
            'updated_by': session['username']
        })
        
        # Invalidate sales cache to reflect changes immediately
        cache['sales_orders']['data'] = None
        
        # Log activity
        activity_data = {
            'action': 'Delivered Status Toggled',
            'details': f'Sale {sale_id} marked as {"delivered" if new_delivered_status else "not delivered"}',
            'user': session['username'],
            'timestamp': datetime.now()
        }
        db.collection('activities').add(activity_data)
        
        return jsonify({
            'success': True, 
            'message': f'Item marked as {"delivered" if new_delivered_status else "not delivered"} successfully!',
            'delivered': new_delivered_status
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error toggling delivered status: {str(e)}'})

# Get Customer History
@app.route('/get_customer_history')
@login_required
@permission_required('sales_customer')
def get_customer_history():
    try:
        customer_name = request.args.get('name', '').strip()
        customer_phone = request.args.get('phone', '').strip()
        
        if not customer_name and not customer_phone:
            return jsonify({'error': 'Name or phone required'})
        
        # Query sales orders for this customer
        sales_ref = db.collection('sales_orders')
        
        # Build query based on available parameters
        if customer_name and customer_phone:
            # Query by both name and phone
            name_query = sales_ref.where('customer_name', '==', customer_name)
            phone_query = sales_ref.where('customer_phone', '==', customer_phone)
            
            # Get results from both queries
            name_results = name_query.stream()
            phone_results = phone_query.stream()
            
            # Combine and deduplicate results
            all_sales = {}
            for doc in name_results:
                all_sales[doc.id] = doc.to_dict()
            for doc in phone_results:
                all_sales[doc.id] = doc.to_dict()
                
        elif customer_name:
            name_query = sales_ref.where('customer_name', '==', customer_name)
            all_sales = {doc.id: doc.to_dict() for doc in name_query.stream()}
            
        elif customer_phone:
            phone_query = sales_ref.where('customer_phone', '==', customer_phone)
            all_sales = {doc.id: doc.to_dict() for doc in phone_query.stream()}
        
        # Process results
        history = []
        total_orders = 0
        total_returned = 0
        last_order_date = None
        
        for sale_id, sale_data in all_sales.items():
            total_orders += 1
            if sale_data.get('status') == 'returned':
                total_returned += 1
            
            order_date = sale_data.get('created_at')
            if order_date and (not last_order_date or order_date > last_order_date):
                last_order_date = order_date
            
            history.append({
                'id': sale_id,
                'product_name': sale_data.get('product_name', ''),
                'product_category': sale_data.get('product_category', ''),
                'product_color': sale_data.get('product_color', ''),
                'product_size': sale_data.get('product_size', ''),
                'quantity': sale_data.get('quantity', 0),
                'total_price': sale_data.get('total_price', 0),
                'status': sale_data.get('status', 'completed'),
                'order_date': order_date.strftime('%Y-%m-%d %H:%M') if order_date else 'N/A',
                'returned_at': sale_data.get('returned_at').strftime('%Y-%m-%d %H:%M') if sale_data.get('returned_at') else None
            })
        
        # Sort by date (newest first)
        history.sort(key=lambda x: x['order_date'], reverse=True)
        
        return jsonify({
            'success': True,
            'total_orders': total_orders,
            'total_returned': total_returned,
            'last_order_date': last_order_date.strftime('%Y-%m-%d') if last_order_date else None,
            'history': history[:10]  # Limit to last 10 orders
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

# Get Returned Customers Data
@app.route('/get_returned_customers')
@login_required
@permission_required('sales_customer')
def get_returned_customers():
    try:
        # Get all returned sales
        sales_ref = db.collection('sales_orders')
        returned_sales = sales_ref.where('status', '==', 'returned').get()
        
        # Group by customer
        returned_customers = {}
        for sale_doc in returned_sales:
            sale_data = sale_doc.to_dict()
            customer_key = f"{sale_data.get('customer_name', '')}|{sale_data.get('customer_phone', '')}"
            
            if customer_key not in returned_customers:
                returned_customers[customer_key] = {
                    'customer_name': sale_data.get('customer_name', ''),
                    'customer_phone': sale_data.get('customer_phone', ''),
                    'customer_address': sale_data.get('customer_address', ''),
                    'total_returned_items': 0,
                    'total_returned_value': 0,
                    'returned_items': [],
                    'first_return_date': None,
                    'last_return_date': None
                }
            
            customer = returned_customers[customer_key]
            customer['total_returned_items'] += 1
            customer['total_returned_value'] += sale_data.get('total_price', 0)
            
            return_date = sale_data.get('returned_at')
            if return_date:
                if not customer['first_return_date'] or return_date < customer['first_return_date']:
                    customer['first_return_date'] = return_date
                if not customer['last_return_date'] or return_date > customer['last_return_date']:
                    customer['last_return_date'] = return_date
            
            customer['returned_items'].append({
                'id': sale_doc.id,
                'product_name': sale_data.get('product_name', ''),
                'product_category': sale_data.get('product_category', ''),
                'product_color': sale_data.get('product_color', ''),
                'product_size': sale_data.get('product_size', ''),
                'quantity': sale_data.get('quantity', 0),
                'item_numbers': sale_data.get('item_numbers', ''),
                'total_price': sale_data.get('total_price', 0),
                'returned_at': return_date,
                'returned_by': sale_data.get('returned_by', ''),
                'original_order_date': sale_data.get('created_at')
            })
        
        # Convert to list and sort by last return date (newest first)
        returned_list = list(returned_customers.values())
        returned_list.sort(key=lambda x: x['last_return_date'] or datetime.min, reverse=True)
        
        return jsonify({
            'success': True,
            'returned_customers': returned_list,
            'total_returned_customers': len(returned_list),
            'total_returned_items': sum(customer['total_returned_items'] for customer in returned_list),
            'total_returned_value': sum(customer['total_returned_value'] for customer in returned_list)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

# Search Customers for Autocomplete
@app.route('/search_customers')
@login_required
@permission_required('sales_customer')
def search_customers():
    try:
        query = request.args.get('q', '').strip().lower()
        
        if len(query) < 2:  # Minimum 2 characters
            return jsonify({'customers': []})
        
        # Get all sales orders to extract unique customers
        sales_ref = db.collection('sales_orders')
        sales_orders = sales_ref.get()
        
        # Create a set of unique customers
        customers_set = set()
        for order in sales_orders:
            order_data = order.to_dict()
            customer_name = order_data.get('customer_name', '').strip()
            customer_phone = order_data.get('customer_phone', '').strip()
            
            if customer_name or customer_phone:
                customers_set.add((customer_name, customer_phone))
        
        # Filter customers based on query
        matching_customers = []
        for name, phone in customers_set:
            if (query in name.lower() or query in phone.lower()):
                matching_customers.append({
                    'name': name,
                    'phone': phone
                })
        
        # Sort by name
        matching_customers.sort(key=lambda x: x['name'].lower())
        
        # Limit to 10 results
        return jsonify({'customers': matching_customers[:10]})
        
    except Exception as e:
        return jsonify({'error': str(e)})

# Handle service worker requests to prevent 404 logs
@app.route('/sw.js')
def service_worker():
    return '', 204  # No Content response

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5003))
    app.run(debug=False, host='0.0.0.0', port=port)
